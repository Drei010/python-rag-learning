from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable, List, Optional, Sequence, Tuple

import pandas as pd

from core.config import settings
from services.file_service import get_supported_files


@dataclass(frozen=True)
class SheetClassification:
    file: str
    sheet: str
    classification: str
    reason: str


CellGrid = List[List[Any]]
StyleGrid = List[List[Optional[int]]]
MergedRangeList = List[Tuple[int, int, int, int]]


def get_excel_files() -> List[Path]:
    return sorted(
        path
        for path in get_supported_files()
        if path.suffix.lower() in settings.supported_excel_extensions and path.is_file()
    )


def classify_excel_files() -> List[SheetClassification]:
    classifications = []

    for excel_file in get_excel_files():
        classifications.extend(classify_excel_file(excel_file))

    return classifications


def classify_excel_file(file_path: Path) -> List[SheetClassification]:
    suffix = file_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        classifications = _classify_openpyxl_workbook(file_path)
    else:
        classifications = _classify_pandas_workbook(file_path)

    for result in classifications:
        print(f"{result.file} - {result.sheet}: {result.classification}")

    return classifications


def classify_sheet(
    values: CellGrid,
    styles: Optional[StyleGrid] = None,
    merged_ranges: Optional[MergedRangeList] = None,
) -> Tuple[str, str]:
    bounds = _used_bounds(values)
    if bounds is None:
        return "unstructured", "sheet has no data"

    min_row, max_row, min_col, max_col = bounds
    used_values = [
        row[min_col : max_col + 1] for row in values[min_row : max_row + 1]
    ]
    used_styles = None
    if styles is not None:
        used_styles = [
            row[min_col : max_col + 1] for row in styles[min_row : max_row + 1]
        ]

    # --- Check: merged cells (NEW) ---
    if merged_ranges and _has_significant_merged_cells(merged_ranges, bounds):
        return "unstructured", "sheet contains merged cell regions"

    # --- Check: blank row inside used range (existing) ---
    if any(all(_is_blank(value) for value in row) for row in used_values):
        return "unstructured", "blank row inside used range"

    # --- Check: rows have different widths (existing) ---
    column_count = max_col - min_col + 1
    if any(len(row) != column_count for row in used_values):
        return "unstructured", "rows have different widths"

    # --- Check: header row contains blank cells (existing) ---
    if len(used_values) > 1 and any(_is_blank(value) for value in used_values[0]):
        return "unstructured", "header row contains blank cells"

    # --- Check: sparsity (NEW) ---
    if len(used_values) > 3 and _is_too_sparse(used_values):
        return "unstructured", "data is too sparse"

    # --- Check: multi-row headers (NEW) ---
    if _has_multi_row_headers(used_values):
        return "unstructured", "multiple header rows suggest hierarchical layout"

    # --- Check: columns have mixed types (existing) ---
    if not _columns_have_consistent_types(used_values):
        return "unstructured", "columns contain mixed data types"

    # --- Check: subtotal rows within data (NEW) ---
    if _has_subtotal_rows(used_values):
        return "unstructured", "subtotal rows found within data"

    # --- Check: repeating group labels (NEW) ---
    if _has_repeating_group_labels(used_values):
        return "unstructured", "column contains repeating group labels"

    # --- Check: styles not uniform (existing) ---
    if used_styles is not None and not _styles_are_uniform(used_styles):
        return "unstructured", "cell styles are not uniform"

    return "structured", "rows and columns are uniform"


# ---------------------------------------------------------------------------
# Openpyxl and pandas workbook classification
# ---------------------------------------------------------------------------


def _classify_openpyxl_workbook(file_path: Path) -> List[SheetClassification]:
    from openpyxl import load_workbook

    workbook = load_workbook(file_path, read_only=False, data_only=True)
    classifications = []

    try:
        for worksheet in workbook.worksheets:
            values = [
                [cell.value for cell in row]
                for row in worksheet.iter_rows()
            ]
            styles = [
                [cell.style_id for cell in row]
                for row in worksheet.iter_rows()
            ]
            merged_ranges = [
                (
                    merged.min_row - 1,
                    merged.max_row - 1,
                    merged.min_col - 1,
                    merged.max_col - 1,
                )
                for merged in worksheet.merged_cells.ranges
            ]
            classification, reason = classify_sheet(values, styles, merged_ranges)
            classifications.append(
                SheetClassification(
                    file=file_path.name,
                    sheet=worksheet.title,
                    classification=classification,
                    reason=reason,
                )
            )
    finally:
        workbook.close()

    return classifications


def _classify_pandas_workbook(file_path: Path) -> List[SheetClassification]:
    sheets = pd.read_excel(file_path, sheet_name=None, header=None)
    classifications = []

    for sheet_name, dataframe in sheets.items():
        values = dataframe.where(pd.notna(dataframe), None).values.tolist()
        classification, reason = classify_sheet(values)
        classifications.append(
            SheetClassification(
                file=file_path.name,
                sheet=str(sheet_name),
                classification=classification,
                reason=reason,
            )
        )

    return classifications


# ---------------------------------------------------------------------------
# NEW heuristic: merged cells
# ---------------------------------------------------------------------------


def _has_significant_merged_cells(
    merged_ranges: MergedRangeList,
    bounds: Tuple[int, int, int, int],
) -> bool:
    min_row, max_row, min_col, max_col = bounds

    for mr_min_row, mr_max_row, mr_min_col, mr_max_col in merged_ranges:
        if mr_max_row < min_row or mr_min_row > max_row:
            continue
        if mr_max_col < min_col or mr_min_col > max_col:
            continue

        row_span = mr_max_row - mr_min_row + 1
        col_span = mr_max_col - mr_min_col + 1

        if row_span > 1 or col_span > 3:
            return True

    return False


# ---------------------------------------------------------------------------
# NEW heuristic: sparsity
# ---------------------------------------------------------------------------


def _is_too_sparse(used_values: CellGrid, threshold: float = 0.5) -> bool:
    total_cells = 0
    blank_cells = 0

    for row in used_values:
        for value in row:
            total_cells += 1
            if _is_blank(value):
                blank_cells += 1

    if total_cells == 0:
        return False

    return (blank_cells / total_cells) > threshold


# ---------------------------------------------------------------------------
# NEW heuristic: multi-row headers
# ---------------------------------------------------------------------------


def _has_multi_row_headers(used_values: CellGrid) -> bool:
    if len(used_values) < 3:
        return False

    consecutive_text_rows = 0
    for row in used_values:
        non_blank_values = [v for v in row if not _is_blank(v)]
        if not non_blank_values:
            break
        if all(_cell_type(v) == "text" for v in non_blank_values):
            consecutive_text_rows += 1
        else:
            break

    if consecutive_text_rows < 2:
        return False

    remaining_rows = used_values[consecutive_text_rows:]
    for row in remaining_rows:
        for value in row:
            if not _is_blank(value) and _cell_type(value) in ("number", "datetime"):
                return True

    return False


# ---------------------------------------------------------------------------
# NEW heuristic: subtotal rows
# ---------------------------------------------------------------------------

_SUBTOTAL_KEYWORDS = {"total", "subtotal", "sum", "grand total", "sub-total"}


def _has_subtotal_rows(used_values: CellGrid) -> bool:
    if len(used_values) < 3:
        return False

    header_offset = (
        1
        if _looks_like_header(used_values[0], used_values[1:])
        else 0
    )
    data_rows = used_values[header_offset:]

    if len(data_rows) < 2:
        return False

    for row in data_rows[:-1]:
        if _row_contains_subtotal_keyword(row):
            return True

    return False


def _row_contains_subtotal_keyword(row: List[Any]) -> bool:
    for value in row:
        if not isinstance(value, str):
            continue
        normalized = value.strip().lower()
        if any(keyword in normalized for keyword in _SUBTOTAL_KEYWORDS):
            return True
    return False


# ---------------------------------------------------------------------------
# NEW heuristic: repeating group labels
# ---------------------------------------------------------------------------


def _has_repeating_group_labels(used_values: CellGrid) -> bool:
    if len(used_values) < 5:
        return False

    header_offset = (
        1
        if _looks_like_header(used_values[0], used_values[1:])
        else 0
    )
    data_rows = used_values[header_offset:]

    if len(data_rows) < 5:
        return False

    column_count = len(data_rows[0])

    for col_index in range(column_count):
        column_values = [
            row[col_index] if col_index < len(row) else None
            for row in data_rows
        ]
        longest_run = _longest_consecutive_run(column_values)
        if longest_run >= 4 and longest_run / len(data_rows) > 0.3:
            return True

    return False


def _longest_consecutive_run(values: List[Any]) -> int:
    max_run = 0
    current_run = 0
    previous_value = object()

    for value in values:
        if _is_blank(value):
            current_run = 0
            previous_value = object()
            continue

        if value == previous_value:
            current_run += 1
        else:
            current_run = 1
            previous_value = value

        if current_run > max_run:
            max_run = current_run

    return max_run


# ---------------------------------------------------------------------------
# Existing helper functions
# ---------------------------------------------------------------------------


def _used_bounds(values: CellGrid) -> Optional[Tuple[int, int, int, int]]:
    non_blank_cells = [
        (row_index, column_index)
        for row_index, row in enumerate(values)
        for column_index, value in enumerate(row)
        if not _is_blank(value)
    ]

    if not non_blank_cells:
        return None

    rows = [row_index for row_index, _ in non_blank_cells]
    columns = [column_index for _, column_index in non_blank_cells]
    return min(rows), max(rows), min(columns), max(columns)


def _is_blank(value: Any) -> bool:
    if value is None:
        return True

    try:
        if pd.isna(value):
            return True
    except (TypeError, ValueError):
        pass

    return isinstance(value, str) and not value.strip()


def _columns_have_consistent_types(values: CellGrid) -> bool:
    header_offset = (
        1
        if len(values) > 1 and _looks_like_header(values[0], values[1:])
        else 0
    )
    data_rows = values[header_offset:] or values
    column_count = len(values[0])

    for column_index in range(column_count):
        column_types = {
            _cell_type(row[column_index])
            for row in data_rows
            if not _is_blank(row[column_index])
        }
        if len(column_types) > 1:
            return False

    return True


def _looks_like_header(header: Sequence[Any], data_rows: Sequence[Sequence[Any]]) -> bool:
    if not header or not data_rows:
        return False

    if not all(isinstance(value, str) and value.strip() for value in header):
        return False

    data_column_types = []
    for column_index in range(len(header)):
        types = {
            _cell_type(row[column_index])
            for row in data_rows
            if not _is_blank(row[column_index])
        }
        data_column_types.append(types)

    return any(types and types != {"text"} for types in data_column_types)


def _cell_type(value: Any) -> str:
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (datetime, date, time)):
        return "datetime"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return "number"
    return "text"


def _styles_are_uniform(styles: StyleGrid) -> bool:
    if len(styles) <= 1:
        return True

    body_styles = styles[1:] if _first_row_is_distinct(styles) else styles
    return len({_row_style_signature(row) for row in body_styles}) == 1


def _first_row_is_distinct(styles: StyleGrid) -> bool:
    first_signature = _row_style_signature(styles[0])
    remaining_signatures = {_row_style_signature(row) for row in styles[1:]}
    return len(remaining_signatures) == 1 and first_signature not in remaining_signatures


def _row_style_signature(row: Iterable[Optional[int]]) -> Tuple[Optional[int], ...]:
    return tuple(row)
